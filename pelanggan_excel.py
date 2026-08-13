"""
=====================================
MAP PERTAMINA - MANAJEMEN PELANGGAN DARI EXCEL
=====================================
Membaca data NIK pelanggan dari file Excel.

Kolom Excel (Sheet1): No. | Nama | NIK KTP | Keterangan | ...
Keterangan: 'RT', 'UM', atau 'RT/UM'.

Dipakai oleh core/session_pool.py:
    from pelanggan_excel import PelangganManager
    mgr = PelangganManager(excel_path=...)
    mgr.semua_pelanggan  # → [{nik, nama, keterangan}, ...]
"""

import json
import os
import random
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = "Data_NIK_Konsumen_LPG.xlsx"
LOG_PATH   = "log_transaksi_harian.json"

TABUNG_RT = 1
TABUNG_UM = 3


class PelangganManager:
    """Mengelola data pelanggan dari Excel."""

    def __init__(self, excel_path: str = EXCEL_PATH, sheet_name: str = "Sheet1"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.semua_pelanggan: list[dict] = []
        self.log_hari_ini: dict = {}
        self._load_excel()
        self._load_log()

    # ── LOAD ──
    def _load_excel(self):
        if not Path(self.excel_path).exists():
            raise FileNotFoundError(
                f"File Excel tidak ditemukan: {self.excel_path}\n"
                f"Letakkan file di folder: {os.getcwd()}"
            )

        print(f"  [Excel] Membaca data dari: {self.excel_path}")
        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        self.semua_pelanggan = []
        if not rows:
            print("  [Excel] Kosong"); return

        # ── Petakan kolom BERDASARKAN HEADER (tahan walau urutan kolom beda) ──
        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def cari(*keys, default=None):
            for i, h in enumerate(header):
                if any(k in h for k in keys):
                    return i
            return default

        i_nama = cari("nama", default=1)
        i_nik  = cari("nik",  default=2)
        i_ket  = cari("keterangan", default=3)
        i_tmp  = cari("tempat")
        i_tgl  = cari("tanggal", "lahir")

        try:
            from core.nik_util import parse_tanggal_lahir
        except Exception:
            parse_tanggal_lahir = None

        def ambil(row, idx):
            return row[idx] if (idx is not None and idx < len(row)) else None

        mati_count = 0
        for row in rows[1:]:
            if not row:
                continue
            nik_raw = ambil(row, i_nik)
            if not nik_raw:
                continue
            nik = str(nik_raw).strip()
            if len(nik) != 16 or not nik.isdigit():
                continue

            nama = ambil(row, i_nama)
            nama = str(nama).strip().rstrip("/") if nama else ""
            ket  = ambil(row, i_ket)
            ket  = str(ket).strip().upper() if ket else "RT"
            if ket not in ("RT", "UM", "RT/UM"):
                ket = "RT"

            tempat = ambil(row, i_tmp)
            tempat = str(tempat).strip() if tempat is not None else ""
            tgl_raw = ambil(row, i_tgl)
            if parse_tanggal_lahir:
                tgl, mati = parse_tanggal_lahir(tgl_raw, tempat)
            else:
                tgl, mati = None, ("mati" in tempat.lower())
            if mati:
                mati_count += 1

            self.semua_pelanggan.append({
                "nik":          nik,
                "nama":         nama,
                "keterangan":   ket,
                "tempat_lahir": "" if mati else tempat,
                "tgl_lahir":    tgl,     # (d,m,y) atau None
                "mati":         mati,
            })

        print(f"  [Excel] {len(self.semua_pelanggan)} pelanggan dimuat "
              f"(Tempat: kol {i_tmp}, Tgl: kol {i_tgl}, Mati: {mati_count})")
        rt    = sum(1 for p in self.semua_pelanggan if p["keterangan"] == "RT")
        um    = sum(1 for p in self.semua_pelanggan if p["keterangan"] == "UM")
        rt_um = sum(1 for p in self.semua_pelanggan if p["keterangan"] == "RT/UM")
        print(f"  [Excel] RT: {rt} | UM: {um} | RT/UM: {rt_um}")

    def _load_log(self):
        hari_ini = date.today().isoformat()
        if Path(LOG_PATH).exists():
            try:
                with open(LOG_PATH, "r", encoding="utf-8") as f:
                    semua = json.load(f)
                self.log_hari_ini = semua.get(hari_ini, {})
            except Exception:
                self.log_hari_ini = {}
        else:
            self.log_hari_ini = {}

    def _simpan_log(self):
        hari_ini = date.today().isoformat()
        semua = {}
        if Path(LOG_PATH).exists():
            try:
                with open(LOG_PATH, "r", encoding="utf-8") as f:
                    semua = json.load(f)
            except Exception:
                pass
        semua[hari_ini] = self.log_hari_ini
        with open(LOG_PATH, "w", encoding="utf-8") as f:
            json.dump(semua, f, indent=2, ensure_ascii=False)

    # ── UTIL (opsional, dari versi standalone) ──
    def catat_hasil(self, nik, nama, kategori, jumlah_tabung, status):
        self.log_hari_ini[nik] = {
            "nama": nama, "kategori": kategori, "jumlah_tabung": jumlah_tabung,
            "status": status, "waktu": datetime.now().strftime("%H:%M:%S"),
        }
        self._simpan_log()


# ─────────────────────────────────────────────
# TAMBAH PELANGGAN KE EXCEL
# ─────────────────────────────────────────────

def tambah_pelanggan_ke_excel(excel_path: str, nik: str, nama: str,
                              keterangan: str = "RT") -> None:
    """
    Tambahkan 1 pelanggan (baris baru) ke file Excel.

    Kolom: No. | Nama | NIK KTP | Keterangan  (sheet "Sheet1").
    Validasi: NIK 16 digit angka & belum ada. Raise ValueError jika gagal.
    """
    from openpyxl import load_workbook, Workbook

    nik = str(nik).strip()
    if len(nik) != 16 or not nik.isdigit():
        raise ValueError("NIK harus tepat 16 digit angka.")
    nama = (nama or "").strip()
    if not nama:
        raise ValueError("Nama tidak boleh kosong.")
    keterangan = (keterangan or "RT").strip().upper()
    if keterangan not in ("RT", "UM", "RT/UM"):
        keterangan = "RT"

    p = Path(excel_path)
    if p.exists():
        wb = load_workbook(excel_path)
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["No.", "Nama", "NIK KTP", "Keterangan"])

    max_no = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        if row[2] is not None and str(row[2]).strip() == nik:
            raise ValueError(f"NIK {nik} sudah terdaftar di Excel.")
        if row[0] is not None and str(row[0]).strip().isdigit():
            max_no = max(max_no, int(row[0]))

    ws.append([max_no + 1, nama, nik, keterangan])
    wb.save(excel_path)
    print(f"[Excel] Pelanggan ditambah: {nama} ({nik}) {keterangan} → {excel_path}")


if __name__ == "__main__":
    try:
        mgr = PelangganManager()
        print(f"Total: {len(mgr.semua_pelanggan)} pelanggan")
    except FileNotFoundError as e:
        print(f"❌ {e}")
